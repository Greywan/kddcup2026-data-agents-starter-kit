#!/usr/bin/env python3
"""Collect all task summaries from data/public/input/ into an Excel file."""

import json
from pathlib import Path

INPUT_DIR = Path(__file__).parent.parent / "data" / "public" / "input"
OUTPUT_FILE = Path(__file__).parent.parent / "tasks_summary.xlsx"


def collect_tasks():
    rows = []
    for task_dir in sorted(INPUT_DIR.glob("task_*"), key=lambda p: int(p.name.split("_")[1])):
        task_file = task_dir / "task.json"
        if not task_file.exists():
            continue
        with open(task_file, encoding="utf-8") as f:
            data = json.load(f)
        rows.append({
            "task_id": data["task_id"],
            "difficulty": data["difficulty"],
            "question": data["question"],
        })

    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tasks Summary"
        # Header
        for col_idx, key in enumerate(["task_id", "difficulty", "question"], 1):
            ws.cell(row=1, column=col_idx, value=key)
        # Data
        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row["task_id"])
            ws.cell(row=row_idx, column=2, value=row["difficulty"])
            ws.cell(row=row_idx, column=3, value=row["question"])
        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)
        wb.save(OUTPUT_FILE)
        print(f"Wrote {len(rows)} tasks to {OUTPUT_FILE}")
    except ImportError:
        # Fallback: write CSV
        csv_file = OUTPUT_FILE.with_suffix(".csv")
        with open(csv_file, "w", encoding="utf-8") as f:
            f.write("task_id,difficulty,question\n")
            for row in rows:
                q = row["question"].replace('"', '""')
                f.write(f'{row["task_id"]},{row["difficulty"]},"{q}"\n')
            print(f"openpyxl not available, wrote {len(rows)} tasks to {csv_file}")


if __name__ == "__main__":
    collect_tasks()
