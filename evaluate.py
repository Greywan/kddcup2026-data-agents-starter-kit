#!/usr/bin/env python3
"""Evaluate predictions against gold standard CSVs and generate an Excel report.

Usage:
    python evaluate.py --run-dir artifacts/runs/<run_id> --gold-dir data/public/output
    python evaluate.py --run-dir artifacts/runs/<run_id> --gold-dir data/public/output --task task_11
    python evaluate.py --run-dir artifacts/runs/<run_id> --gold-dir data/public/output --output-report my_report.xlsx
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install it with: uv add openpyxl", file=sys.stderr)
    sys.exit(1)


def load_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    """Load a CSV file, return (columns, rows)."""
    with path.open("r", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return [], []
    columns = rows[0]
    data_rows = rows[1:]
    return columns, data_rows


def evaluate_task(pred_path: Path, gold_path: Path) -> tuple[bool, str]:
    """Evaluate a single task.

    Rules:
    - prediction must contain ALL gold columns (extra columns are OK)
    - values in the matched columns must be exactly the same
    - otherwise score = 0

    Returns (is_correct, detail_message).
    """
    gold_columns, gold_rows = load_csv(gold_path)
    pred_columns, pred_rows = load_csv(pred_path)

    gold_col_set = set(gold_columns)
    pred_col_set = set(pred_columns)

    # Check: prediction must contain ALL gold columns
    missing_cols = gold_col_set - pred_col_set
    if missing_cols:
        return False, f"Missing gold columns: {sorted(missing_cols)}"

    if not gold_columns:
        return True, "Empty gold table — trivially correct"

    # Build column index maps
    pred_col_index = {col: i for i, col in enumerate(pred_columns)}
    gold_col_index = {col: i for i, col in enumerate(gold_columns)}

    # Row count must match
    if len(pred_rows) != len(gold_rows):
        return (
            False,
            f"Row count mismatch: pred={len(pred_rows)}, gold={len(gold_rows)}",
        )

    # Check every cell in gold columns
    mismatched_examples = []
    for row_idx in range(len(gold_rows)):
        for gold_col in gold_columns:
            g_val = gold_rows[row_idx][gold_col_index[gold_col]]
            p_val = pred_rows[row_idx][pred_col_index[gold_col]]
            if g_val != p_val:
                mismatched_examples.append(
                    f"row={row_idx} col={gold_col}: pred='{p_val}' vs gold='{g_val}'"
                )
                if len(mismatched_examples) >= 5:
                    break
        if len(mismatched_examples) >= 5:
            break

    if mismatched_examples:
        detail = "; ".join(mismatched_examples)
        if len(gold_rows) * len(gold_columns) > 5:
            detail += "; ..."
        return False, f"Value mismatch: {detail}"

    return True, "All gold columns match"


def write_excel_report(results: list[dict], output_path: Path, run_name: str) -> None:
    """Generate a formatted Excel report."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluation Results"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    skip_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ok_font = Font(color="006100", bold=True)
    fail_font = Font(color="9C0006", bold=True)
    skip_font = Font(color="9C6500", bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Headers
    headers = ["Task ID", "Status", "Score", "Detail"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Data rows
    for row_idx, result in enumerate(results, 2):
        status = result["status"]
        score = 1 if status == "OK" else 0

        # Task ID
        cell = ws.cell(row=row_idx, column=1, value=result["task_id"])
        cell.border = thin_border
        cell.alignment = center

        # Status
        cell = ws.cell(row=row_idx, column=2, value="PASS" if status == "OK" else ("SKIP" if status == "SKIP" else "FAIL"))
        cell.border = thin_border
        cell.alignment = center
        if status == "OK":
            cell.fill = ok_fill
            cell.font = ok_font
        elif status == "SKIP":
            cell.fill = skip_fill
            cell.font = skip_font
        else:
            cell.fill = fail_fill
            cell.font = fail_font

        # Score
        cell = ws.cell(row=row_idx, column=3, value=score)
        cell.border = thin_border
        cell.alignment = center

        # Detail
        cell = ws.cell(row=row_idx, column=4, value=result["detail"])
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Summary row
    total = sum(1 for r in results if r["status"] != "SKIP")
    correct = sum(1 for r in results if r["status"] == "OK")
    skipped = sum(1 for r in results if r["status"] == "SKIP")
    accuracy = correct / total if total > 0 else 0

    summary_row = len(results) + 2
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=f"{correct}/{total}").font = Font(bold=True)
    ws.cell(row=summary_row, column=3, value=accuracy).font = Font(bold=True)
    ws.cell(row=summary_row, column=3).number_format = "0.0%"
    ws.cell(row=summary_row, column=4, value=f"Skipped: {skipped} (no prediction.csv)").font = Font(bold=True, italic=True)

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 80

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Evaluate predictions against gold CSVs")
    parser.add_argument(
        "--run-dir",
        type=Path,
        required=True,
        help="Path to the run output directory, e.g. artifacts/runs/<run_id>",
    )
    parser.add_argument(
        "--gold-dir",
        type=Path,
        default=Path("data/public/output"),
        help="Path to the gold output directory (default: data/public/output)",
    )
    parser.add_argument(
        "--task",
        type=str,
        default=None,
        help="Evaluate a single task only, e.g. task_11",
    )
    parser.add_argument(
        "--output-report",
        type=Path,
        default=None,
        help="Path to output Excel report (default: <run-dir>/evaluation_report.xlsx)",
    )
    args = parser.parse_args()

    run_dir: Path = args.run_dir
    gold_dir: Path = args.gold_dir

    if not run_dir.is_dir():
        print(f"Error: run-dir not found: {run_dir}", file=sys.stderr)
        sys.exit(1)
    if not gold_dir.is_dir():
        print(f"Error: gold-dir not found: {gold_dir}", file=sys.stderr)
        sys.exit(1)

    # Collect task IDs to evaluate
    if args.task:
        task_ids = [args.task]
    else:
        task_ids = sorted(
            d.name
            for d in gold_dir.iterdir()
            if d.is_dir() and d.name.startswith("task_")
        )
        # Sort by numeric ID (task_11 < task_194, not lexicographic)
        task_ids.sort(key=lambda name: int(name.split("_")[1]))

    results = []

    for task_id in task_ids:
        pred_path = run_dir / task_id / "prediction.csv"
        gold_path = gold_dir / task_id / "gold.csv"

        if not gold_path.exists():
            results.append({
                "task_id": task_id,
                "status": "SKIP",
                "detail": "gold.csv not found",
            })
            continue

        if not pred_path.exists():
            results.append({
                "task_id": task_id,
                "status": "FAIL",
                "detail": "prediction.csv not found",
            })
            continue

        is_correct, detail = evaluate_task(pred_path, gold_path)
        results.append({
            "task_id": task_id,
            "status": "OK" if is_correct else "FAIL",
            "detail": detail,
        })

    # Console output
    for result in results:
        status_icon = {"OK": "  OK  ", "FAIL": "  FAIL", "SKIP": "  SKIP"}.get(result["status"], "")
        if result["status"] == "OK":
            print(f"{status_icon} {result['task_id']}")
        else:
            print(f"{status_icon} {result['task_id']}: {result['detail']}")

    # Summary
    total = sum(1 for r in results if r["status"] != "SKIP")
    correct = sum(1 for r in results if r["status"] == "OK")
    skipped = sum(1 for r in results if r["status"] == "SKIP")
    accuracy = correct / total if total > 0 else 0

    print()
    print(f"Results: {correct}/{total} correct ({accuracy:.1%})")
    if skipped:
        print(f"Skipped: {skipped} (no gold.csv)")

    # Generate Excel report
    output_report = args.output_report
    if output_report is None:
        output_report = run_dir / "evaluation_report.xlsx"

    write_excel_report(results, output_report, run_dir.name)


if __name__ == "__main__":
    main()
